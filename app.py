import hashlib
import io
import logging
import os
import sys
import threading
import zipfile
from datetime import datetime

import pandas as pd
import streamlit as st

# Silence fontTools "MERG NOT subset" console noise during PDF font subsetting
logging.getLogger("fontTools").setLevel(logging.ERROR)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from fpdf import FPDF
    from fpdf.fonts import FontFace

    HAS_FPDF = True
except ImportError:
    HAS_FPDF = False

# Force light theme even if Streamlit was started outside the project folder
try:
    from streamlit import config as _st_config

    _st_config.set_option("theme.base", "light")
    _st_config.set_option("theme.primaryColor", "#31869B")
    _st_config.set_option("theme.backgroundColor", "#FFFFFF")
    _st_config.set_option("theme.secondaryBackgroundColor", "#F2F8FA")
    _st_config.set_option("theme.textColor", "#1F2937")
except Exception:
    pass

st.set_page_config(page_title="Arji Report Splitter", page_icon="📋", layout="wide")

# ---------- White-mode styling ----------
st.markdown(
    """
    <style>
    .stApp { background-color: #FFFFFF; }

    /* Force dark readable text everywhere (even if browser is in dark mode) */
    .stApp h1, .stApp h2, .stApp h4, .stApp h5 { color: #1F2937; }
    [data-testid="stWidgetLabel"] p, [data-testid="stWidgetLabel"] label,
    [data-testid="stMarkdownContainer"] p, .stCheckbox p { color: #1F2937 !important; }
    [data-testid="stCaptionContainer"] p { color: #6B7280 !important; }

    /* Header banner */
    .app-header {
        background: linear-gradient(90deg, #31869B 0%, #92CDDC 100%);
        border-radius: 12px;
        padding: 22px 30px;
        margin-bottom: 18px;
    }
    .app-header h1 { color: #FFFFFF !important; margin: 0; font-size: 1.9rem; }
    .app-header p  { color: #EAF6FA !important; margin: 6px 0 0 0; font-size: 0.95rem; }

    /* Metric cards */
    div[data-testid="stMetric"] {
        background: #F2F8FA;
        border: 1px solid #92CDDC;
        border-radius: 10px;
        padding: 14px 18px;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.06);
    }
    div[data-testid="stMetric"] label { color: #31869B !important; font-weight: 600; }
    div[data-testid="stMetricValue"] { color: #1F2937 !important; font-weight: 700; }

    /* Section headings */
    .stApp h3 { color: #31869B; border-bottom: 2px solid #92CDDC; padding-bottom: 4px; }

    /* Select / multiselect boxes — light background, blue border */
    div[data-baseweb="select"] > div {
        background-color: #F9FCFD !important;
        border-color: #92CDDC !important;
        color: #1F2937 !important;
    }
    div[data-baseweb="select"] input { color: #1F2937 !important; }
    div[data-baseweb="select"] svg { fill: #31869B !important; }

    /* Dropdown open menu */
    div[data-baseweb="popover"] ul { background-color: #FFFFFF !important; }
    div[data-baseweb="popover"] li { color: #1F2937 !important; background-color: #FFFFFF !important; }
    div[data-baseweb="popover"] li:hover { background-color: #E6F3F7 !important; }

    /* Selected chips/tags — teal instead of red */
    span[data-baseweb="tag"] {
        background-color: #31869B !important;
        color: #FFFFFF !important;
    }
    span[data-baseweb="tag"] span { color: #FFFFFF !important; }
    span[data-baseweb="tag"] svg { fill: #FFFFFF !important; }

    /* Tabs */
    button[data-baseweb="tab"] { font-weight: 600; color: #4B5563 !important; }
    button[data-baseweb="tab"][aria-selected="true"] { color: #31869B !important; }
    div[data-baseweb="tab-highlight"] { background-color: #31869B !important; }

    /* File uploader — light dashed drop zone */
    div[data-testid="stFileUploader"] {
        border: 2px dashed #92CDDC;
        border-radius: 10px;
        padding: 10px;
        background: #F9FCFD;
    }
    [data-testid="stFileUploaderDropzone"] {
        background-color: #F2F8FA !important;
        color: #1F2937 !important;
    }
    [data-testid="stFileUploaderDropzone"] div, [data-testid="stFileUploaderDropzone"] span,
    [data-testid="stFileUploaderDropzone"] small { color: #1F2937 !important; }
    [data-testid="stFileUploaderDropzone"] svg { fill: #31869B !important; }
    [data-testid="stFileUploaderDropzone"] button {
        background-color: #31869B !important;
        color: #FFFFFF !important;
        border: none !important;
        border-radius: 8px;
    }
    [data-testid="stFileUploaderFile"], [data-testid="stFileUploaderFile"] div,
    [data-testid="stFileUploaderFile"] small { color: #1F2937 !important; }
    [data-testid="stFileUploaderFile"] svg { fill: #31869B !important; }

    /* Name-selector popover — looks like a dropdown field */
    div[data-testid="stPopover"] button {
        background-color: #F9FCFD !important;
        border: 1px solid #92CDDC !important;
        color: #1F2937 !important;
        border-radius: 8px;
        width: 100%;
        justify-content: space-between;
        font-weight: 500;
    }
    div[data-testid="stPopover"] button:hover {
        border-color: #31869B !important;
        color: #31869B !important;
    }
    div[data-testid="stPopover"] button p { color: inherit !important; }
    div[data-testid="stPopoverBody"] {
        background-color: #FFFFFF !important;
        border: 1px solid #92CDDC;
        border-radius: 8px;
    }

    /* Expanders (Summary / Preview) — light card with blue border */
    div[data-testid="stExpander"] details {
        border: 1px solid #92CDDC;
        border-radius: 10px;
        background: #F9FCFD;
    }
    div[data-testid="stExpander"] summary { color: #31869B !important; font-weight: 600; }
    div[data-testid="stExpander"] summary p { color: #31869B !important; font-weight: 600; }
    div[data-testid="stExpander"] summary svg { fill: #31869B !important; }

    /* Download buttons — teal, full width */
    div[data-testid="stDownloadButton"] button {
        background-color: #31869B !important;
        color: #FFFFFF !important;
        border: 1px solid #31869B !important;
        border-radius: 8px;
        font-weight: 600;
        width: 100%;
    }
    div[data-testid="stDownloadButton"] button:hover { background-color: #256B7D !important; }
    div[data-testid="stDownloadButton"] button p { color: #FFFFFF !important; }
    </style>
    <div class="app-header">
        <h1>📋 Arji Report Splitter</h1>
        <p>Upload the visitor/arji Excel file, select <b>Meet To</b> name(s) —
        the app splits data into 4 sheets: Selected Pending, Selected Completed, Other Pending, Other Completed.</p>
    </div>
    """,
    unsafe_allow_html=True,
)

MEET_TO_COL = "Meet To"
STATUS_COL = "Status"

# ---------- 2-column layout: upload (left) | name & status selection (right) ----------
left_col, right_col = st.columns(2, gap="large")

with left_col:
    st.markdown("#### 📤 Upload File")
    uploaded = st.file_uploader("Upload Excel file", type=["xlsx", "xls"])

if uploaded is None:
    with right_col:
        st.markdown("#### 🔽 Select Meet To (Name)")
        st.info("👈 Upload the Excel file first — name & status options will appear here.")
    st.stop()

# ---------- Read file (header may be in row 1, 2 or 3) ----------
file_bytes = uploaded.getvalue()


@st.cache_data(show_spinner=False, max_entries=3)
def load_excel(raw: bytes):
    """Parse once per uploaded file — reruns (checkbox clicks) reuse the cache."""
    tried = {}
    for header_row in range(3):  # try first 3 rows as header
        candidate = pd.read_excel(io.BytesIO(raw), header=header_row)
        candidate.columns = [str(c).strip() for c in candidate.columns]
        tried[header_row + 1] = list(candidate.columns)
        if MEET_TO_COL in candidate.columns and STATUS_COL in candidate.columns:
            return candidate, header_row, tried
    return None, -1, tried


try:
    df, header_row_used, tried_columns = load_excel(file_bytes)
except Exception as e:
    st.error(f"Could not read the Excel file: {e}")
    st.stop()

if df is None:
    st.error(
        f"Column(s) '{MEET_TO_COL}' and '{STATUS_COL}' not found in the first 3 rows of the file."
    )
    for row_no, cols in tried_columns.items():
        st.write(f"Row {row_no} headers:", cols)
    st.stop()

with left_col:
    st.success(f"File loaded — {len(df)} rows.")
    if header_row_used > 0:
        st.caption(f"ℹ️ Header found in row {header_row_used + 1} of the file.")

# ---------- Select names + status (right column) ----------
meet_to_values = (
    df[MEET_TO_COL].dropna().astype(str).str.strip().replace("", pd.NA).dropna().unique().tolist()
)
meet_to_values = sorted(meet_to_values)

def _toggle_all_names():
    """When (Select All) is clicked, set every name checkbox to match it."""
    val = st.session_state["select_all_names"]
    for n in meet_to_values:
        st.session_state[f"chk_{n}"] = val


with right_col:
    st.markdown("#### 🔽 Select Meet To (Name)")

    if hasattr(st, "popover"):
        # Nothing selected by default — user picks name(s).
        # Keep (Select All) in sync: unchecking any name unticks it, like Excel
        name_states = [bool(st.session_state.get(f"chk_{n}", False)) for n in meet_to_values]
        if "select_all_names" in st.session_state:
            st.session_state["select_all_names"] = all(name_states)

        sel_count = sum(name_states)
        with st.popover(
            f"👤 Select Meet To (names) — {sel_count} of {len(meet_to_values)} selected",
            width="stretch",
        ):
            st.checkbox(
                "(Select All)",
                value=False,
                key="select_all_names",
                on_change=_toggle_all_names,
            )
            st.divider()
            selected_names = []
            for n in meet_to_values:
                if st.checkbox(n, value=False, key=f"chk_{n}"):
                    selected_names.append(n)
    else:
        # Older Streamlit without st.popover — fall back to multiselect
        selected_names = st.multiselect("Select Meet To (names)", meet_to_values)

if not selected_names:
    st.info("👉 Select name(s) from the dropdown above to continue.")
    st.stop()

# Label for the selected group: single name, or "Selected (N)"
if len(selected_names) == 1:
    group_label = selected_names[0]
else:
    group_label = f"Selected ({len(selected_names)} names)"

# ---------- Split into 4 sets ----------
meet_series = df[MEET_TO_COL].astype(str).str.strip()
status_series = df[STATUS_COL].astype(str).str.strip()

selected_set = {str(n).strip() for n in selected_names}
is_cp = meet_series.isin(selected_set)
# Fixed rule: status containing "pending" = Pending, everything else = Completed
is_pending = status_series.str.lower().str.contains("pend", na=False)

cp_pending = df[is_cp & is_pending]
cp_completed = df[is_cp & ~is_pending]
other_pending = df[~is_cp & is_pending]
other_completed = df[~is_cp & ~is_pending]

# ---------- Summary (hidden — click to open) ----------
with st.expander("📊 Summary — click to view counts", expanded=False):
    m1, m2, m3, m4 = st.columns(4)
    m1.metric(f"{group_label} — Pending", len(cp_pending))
    m2.metric(f"{group_label} — Completed", len(cp_completed))
    m3.metric("Other — Pending", len(other_pending))
    m4.metric("Other — Completed", len(other_completed))

    # Per-name breakdown when more than one name is selected
    if len(selected_names) > 1:
        breakdown = (
            pd.DataFrame(
                {
                    "Meet To": meet_series[is_cp],
                    "Pending": is_pending[is_cp].astype(int),
                    "Completed": (~is_pending[is_cp]).astype(int),
                }
            )
            .groupby("Meet To", as_index=False)
            .sum()
            .sort_values("Meet To")
        )
        breakdown["Total"] = breakdown["Pending"] + breakdown["Completed"]
        st.markdown("**Per-name breakdown (selected names):**")
        st.dataframe(breakdown, width="stretch", hide_index=True)

# ---------- Output format (same as uploaded sheet) ----------
# Selected person sheets: no "Meet To" column
OUTPUT_COLS_SELECTED = [
    "Sr.No.",
    "Police Station",
    "Visitor Name",
    "Visitor Mobile No",
    "Arji No",
    "Visit Purpose",
    "VisitDateTime",
    "Details of Actions Taken (कार्यवाही का विवरण)",
]

# Other sheets: include "Meet To" so you can see whose arji it is
OUTPUT_COLS_OTHER = [
    "Sr.No.",
    "Police Station",
    "Meet To",
    "Visitor Name",
    "Visitor Mobile No",
    "Arji No",
    "Visit Purpose",
    "VisitDateTime",
    "Details of Actions Taken (कार्यवाही का विवरण)",
]

COL_WIDTHS = {
    "Sr.No.": 8,
    "Police Station": 18,
    "Meet To": 20,
    "Visitor Name": 25,
    "Visitor Mobile No": 18,
    "Arji No": 12,
    "Visit Purpose": 30,
    "VisitDateTime": 20,
    "Details of Actions Taken (कार्यवाही का विवरण)": 45,
}

TITLE_PENDING = "Visitor Details(Pending)"
TITLE_COMPLETED = "Visitor Details(Completed)"


def format_output(data: pd.DataFrame, cols: list) -> pd.DataFrame:
    """Keep only the given output columns, renumber Sr.No. from 1."""
    out = pd.DataFrame()
    for col in cols:
        if col == "Sr.No.":
            out[col] = range(1, len(data) + 1)
        elif col in data.columns:
            out[col] = data[col].values
        else:
            out[col] = ""
    return out


# ---------- Downloads ----------
def to_excel_bytes(sheet_dict: dict) -> bytes:
    """sheet_dict: {sheet_name: (title, dataframe, cols)} — writes title row + header + data."""
    buf = io.BytesIO()
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sheet_name, (title, data, cols) in sheet_dict.items():
            out = format_output(data, cols)
            n_cols = len(cols)
            # Excel sheet names max 31 chars; data starts at row 3 (row 1 = title, row 2 = header)
            out.to_excel(writer, sheet_name=sheet_name[:31], index=False, startrow=1)
            ws = writer.sheets[sheet_name[:31]]

            # Title row, merged across all columns — e.g. "Visitor Details(Pending)"
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.font = Font(name="Calibri", bold=True, size=24)
            title_cell.alignment = center
            title_cell.border = border
            ws.row_dimensions[1].height = 32

            # Header row: #92CDDC fill, white bold Calibri 11, centered + wrap
            for c in range(1, n_cols + 1):
                h = ws.cell(row=2, column=c)
                h.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
                h.fill = header_fill
                h.alignment = center
                h.border = border

            # Data cells: Calibri 11, centered + middle + wrap, borders
            for r in range(3, len(out) + 3):
                for c in range(1, n_cols + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.font = Font(name="Calibri", size=11)
                    cell.border = border
                    cell.alignment = center

            # Column widths
            for i, col in enumerate(cols, start=1):
                ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS.get(col, 18)
    return buf.getvalue()


def _pdf_font():
    """Find a Devanagari-capable TTF font (regular, bold) for Hindi text in PDFs.
    Bundled Mukta first — works on any server (incl. Streamlit Cloud/Linux) —
    then Windows system fonts, then common Linux paths."""
    here = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        (
            os.path.join(here, "fonts", "Mukta-Regular.ttf"),
            os.path.join(here, "fonts", "Mukta-Bold.ttf"),
        ),
        (os.path.expandvars(r"%LOCALAPPDATA%\Microsoft\Windows\Fonts\Nirmala.ttf"), None),
        (r"C:\Windows\Fonts\Nirmala.ttf", None),
        (r"C:\Windows\Fonts\mangal.ttf", None),
        ("/usr/share/fonts/truetype/lohit-devanagari/Lohit-Devanagari.ttf", None),
    ]
    for regular, bold in candidates:
        if os.path.exists(regular):
            if not (bold and os.path.exists(bold)):
                bold = regular
            return regular, bold
    return None, None


def _cell_text(v) -> str:
    if pd.isna(v):
        return ""
    if isinstance(v, pd.Timestamp):
        return v.strftime("%d-%m-%Y %H:%M")
    return str(v).replace("\r", "")


def _fit_text(text: str, chars_per_line: int, max_lines: int) -> str:
    """Trim text so its estimated wrapped height fits on one PDF page.
    Accounts for both length and newlines. (Excel keeps the full text.)"""

    def est_lines(t: str) -> int:
        return sum(max(1, -(-len(s) // chars_per_line)) for s in t.split("\n"))

    if est_lines(text) <= max_lines:
        return text
    lo, hi = 0, len(text)
    while lo < hi:
        mid = (lo + hi) // 2
        if est_lines(text[:mid]) <= max_lines - 1:
            lo = mid + 1
        else:
            hi = mid
    return text[: max(0, lo - 1)] + " …"


def to_pdf_bytes(
    title: str, data: pd.DataFrame, cols: list, shaping: bool = True, cap_scale: float = 1.0
) -> bytes:
    """One PDF per sheet: title + #92CDDC header + bordered table (landscape A4)."""
    out = format_output(data, cols)

    # Per-column text budget so no row grows taller than one page.
    # Landscape A4: 297mm wide, 8mm margins → 281mm of table width.
    total_rel = sum(COL_WIDTHS.get(c, 18) for c in cols)
    max_lines = max(5, int(26 * cap_scale))  # 26 lines × 5mm ≈ 130mm, safely < page
    col_chars = {
        c: max(4, int(COL_WIDTHS.get(c, 18) / total_rel * 281 / 3.0)) for c in cols
    }
    pdf = FPDF(orientation="L", format="A4")
    pdf.set_margins(8, 8, 8)
    pdf.set_auto_page_break(auto=True, margin=8)
    pdf.add_page()

    font_regular, font_bold = _pdf_font()
    if font_regular:
        pdf.add_font("Uni", "", font_regular)
        pdf.add_font("Uni", "B", font_bold)
        font = "Uni"
        sanitize = False
        if shaping:
            try:
                pdf.set_text_shaping(True)  # correct Hindi matra/conjunct rendering
            except Exception:
                pass
    else:
        # No Unicode font available — helvetica can't encode Hindi, so replace
        # unsupported characters with '?' instead of crashing.
        font = "helvetica"
        sanitize = True

    def _safe(text: str) -> str:
        if sanitize:
            return text.encode("latin-1", "replace").decode("latin-1")
        return text

    pdf.set_font(font, "B", 16)
    pdf.cell(0, 10, _safe(title), align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    pdf.set_font(font, "", 8)
    header_style = FontFace(
        family=font,
        emphasis="BOLD",
        size_pt=8,
        color=(255, 255, 255),
        fill_color=(146, 205, 220),  # #92CDDC
    )
    widths = tuple(COL_WIDTHS.get(c, 18) for c in cols)
    with pdf.table(
        col_widths=widths,
        text_align="CENTER",
        headings_style=header_style,
        line_height=5,
        padding=1,
    ) as table:
        hrow = table.row()
        for c in cols:
            hrow.cell(_safe(c))
        for _, r in out.iterrows():
            row = table.row()
            for c in cols:
                row.cell(_safe(_fit_text(_cell_text(r[c]), col_chars[c], max_lines)))
    return bytes(pdf.output())


def make_pdf(title: str, data: pd.DataFrame, cols: list) -> bytes:
    """fpdf's bidi algorithm recurses once per character of a long cell, which
    overflows the default recursion limit / thread stack. Render the PDF in a
    dedicated thread with a large stack and raised recursion limit; if a cell
    is still too extreme, retry without text shaping instead of crashing."""
    result = {}

    def _runner():
        attempts = [
            dict(shaping=True, cap_scale=1.0),
            dict(shaping=True, cap_scale=0.5),
            dict(shaping=True, cap_scale=0.25),
            dict(shaping=False, cap_scale=0.25),
        ]
        last_err = None
        for kw in attempts:
            try:
                result["pdf"] = to_pdf_bytes(title, data, cols, **kw)
                return
            except (RecursionError, ValueError) as e:
                last_err = e
            except Exception as e:
                result["err"] = e
                return
        result["err"] = last_err

    old_limit = sys.getrecursionlimit()
    old_stack = threading.stack_size()
    sys.setrecursionlimit(200_000)
    try:
        threading.stack_size(128 * 1024 * 1024)  # 128 MB stack for deep recursion
    except (ValueError, RuntimeError):
        pass
    try:
        t = threading.Thread(target=_runner, daemon=True)
        t.start()
        t.join()
    finally:
        sys.setrecursionlimit(old_limit)
        try:
            threading.stack_size(old_stack)
        except (ValueError, RuntimeError):
            pass

    if "err" in result:
        raise result["err"]
    return result["pdf"]


# Fixed file label: whoever is selected, their sheets are named "CP-Sir"
safe_cp = "CP-Sir"

@st.cache_data(show_spinner=False, max_entries=5)
def build_zip(cp_p, cp_c, ot_p, ot_c, label_key: str, ts: str) -> bytes:
    """Build the ZIP (4 Excel + 4 PDF). Excel and PDF share the same base name,
    e.g. CP-Sir-Completed_30-07-2026_12-10-05.xlsx / .pdf"""
    zip_items = [
        (f"{label_key}-Pending", TITLE_PENDING, cp_p, OUTPUT_COLS_SELECTED),
        (f"{label_key}-Completed", TITLE_COMPLETED, cp_c, OUTPUT_COLS_SELECTED),
        ("Other-Pending", TITLE_PENDING, ot_p, OUTPUT_COLS_OTHER),
        ("Other-Completed", TITLE_COMPLETED, ot_c, OUTPUT_COLS_OTHER),
    ]
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, title, data, cols in zip_items:
            base = f"{name}_{ts}"
            zf.writestr(f"Excel/{base}.xlsx", to_excel_bytes({name[:31]: (title, data, cols)}))
            if HAS_FPDF:
                zf.writestr(f"PDF/{base}.pdf", make_pdf(title, data, cols))
    return zip_buf.getvalue()


st.subheader("Download")

if not HAS_FPDF:
    st.warning(
        "PDF library not installed — ZIP will contain only the 4 Excel files. "
        "Run: pip install fpdf2 uharfbuzz"
    )
elif not _pdf_font()[0]:
    st.warning(
        "Hindi font not found on this server — PDFs will show '?' for Hindi text. "
        "Deploy the app together with its `fonts/` folder (Mukta-Regular.ttf, Mukta-Bold.ttf)."
    )

# Files are built only when Generate is clicked — name selection stays instant.
sel_key = hashlib.md5(file_bytes).hexdigest() + "|" + "|".join(sorted(selected_names))

if st.session_state.get("zip_key") == sel_key:
    st.download_button(
        f"⬇️ Download {st.session_state['zip_name']} (4 Excel + 4 PDF)",
        data=st.session_state["zip_data"],
        file_name=st.session_state["zip_name"],
        mime="application/zip",
        type="primary",
    )
    st.caption("Changing the name selection will need a new Generate.")
else:
    if st.button("🔄 Generate files (4 Excel + 4 PDF)", type="primary"):
        with st.spinner("Preparing ZIP (4 Excel + 4 PDF)..."):
            ts = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
            st.session_state["zip_data"] = build_zip(
                cp_pending, cp_completed, other_pending, other_completed, safe_cp, ts
            )
            st.session_state["zip_key"] = sel_key
            st.session_state["zip_name"] = f"VMS-{ts}.zip"
        st.rerun()
    st.caption("Choose names above, then click **Generate files** to prepare the ZIP.")

# ---------- Preview (hidden — click to open) ----------
with st.expander("👁 Preview sheets — click to view data", expanded=False):
    sheets = {
        f"{group_label} Pending": (cp_pending, OUTPUT_COLS_SELECTED),
        f"{group_label} Completed": (cp_completed, OUTPUT_COLS_SELECTED),
        "Other Pending": (other_pending, OUTPUT_COLS_OTHER),
        "Other Completed": (other_completed, OUTPUT_COLS_OTHER),
    }

    tabs = st.tabs([f"{name} ({len(data)})" for name, (data, _) in sheets.items()])
    for tab, (name, (data, cols)) in zip(tabs, sheets.items()):
        with tab:
            st.dataframe(format_output(data, cols), width="stretch", hide_index=True)
